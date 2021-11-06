#!/usr/bin/env python
from openpyxl import Workbook, load_workbook
import json

class AosXlsx:
    def __init__(self, excel_file, yaml_file) -> None:
        self.excel_file = excel_file
        self.yaml_file = yaml_file
        self.wb = load_workbook(excel_file)
        # captuer whole data here
        self.inventory = {"aos_server": {}, "blueprints": {}}


    def read_header(self, sheet_name, expected):
        print(f"==== reading {sheet_name}")
        print(f"expected: {str(expected)}")
        ws = self.wb[sheet_name]
        headers = dict(((r.value, r.column) for r in ws[1]))
        print(f"given: {headers}")
        # expected = {'host': 2, 'port': 2, 'username': 3, 'password': 4 }
        if not all(h in headers for h in expected):
            print( f"==== incomplete headers: expected {[k for k in expected.keys()]}")
        print(f"return headers: {headers}")
        return (ws, headers)


    def read_aos_server(self):
        sheet_name = "aos_server"
        expected = {'host': 2, 'port': 2, 'username': 3, 'password': 4 }
        (ws, headers) = self.read_header(sheet_name, expected)

        for k in headers:
            self.inventory['aos_server'][k] = ws.cell(row=2, column=headers[k]).value


    def read_blueprints(self):
        sheet_name = "blueprints"
        expected = {'blueprint': 1 }
        (ws, headers) = self.read_header(sheet_name, expected)

        for row_num in range(2, ws.max_row+1):
            my_id = ''
            this_blueprint = {"routing_zones": {}, "systems": {}}
            for header in headers:
                if header == "blueprint":
                    my_id = ws.cell(row=row_num, column=headers[header]).value                
                else:
                    this_blueprint[header] = ws.cell(row=row_num, column=headers[header]).value
            # skip empty row
            if my_id:
                self.inventory["blueprints"][my_id]= this_blueprint


    def read_routing_zones(self):
        sheet_name = "routing_zones"
        expected = {'blueprint': 1, 'routing_zone': 2, 'vrf_name': 3, 'vlan_id': 4, 'sz_type': 5, 'dhcp_servers': 6, 'leaf_loopback_ips': 7, 'to_generic_link_ips': 8 }
        (ws, headers) = self.read_header(sheet_name, expected)

        for row_num in range(2, ws.max_row+1):
            my_id = ''
            this_blueprint = ''
            this_routing_zone = {"virtual_networks": {}}
            for header in headers:
                if header == "blueprint":
                    this_blueprint = ws.cell(row=row_num, column=headers[header]).value                
                elif header == "routing_zone":
                    my_id = ws.cell(row=row_num, column=headers[header]).value                
                    this_routing_zone["label"] = ws.cell(row=row_num, column=headers[header]).value
                elif header in ["dhcp_servers", "leaf_loopback_ips", "to_generic_link_ips" ]:
                    this_routing_zone[header] = json.loads(ws.cell(row=row_num, column=headers[header]).value)
                else:
                    this_routing_zone[header] = ws.cell(row=row_num, column=headers[header]).value
            # skip empty row
            if my_id:
                self.inventory["blueprints"][this_blueprint]["routing_zones"][my_id]= this_routing_zone


    def read_virtual_networks(self):
        sheet_name = "virtual_networks"
        expected = {'blueprint': 1, 'routing_zone': 2, 'virtual_network': 3, 'vrf_name': 4, 'description': 5, 'vn_type': 6, 'vlan_id': 7, 'bound_to': 8, 'virtual_gateway_ipv4_enabled': 9, 'ipv4_enabled': 10, 'ipv4_subnet': 11, 'dhcp_server': 12 }
        (ws, headers) = self.read_header(sheet_name, expected)

        for row_num in range(2, ws.max_row+1):
            my_id = ''
            this_blueprint = ''
            this_routing_zone = ''
            this_virtual_network = {}
            for header in headers:
                if header == "blueprint":
                    this_blueprint = ws.cell(row=row_num, column=headers[header]).value                
                elif header == "routing_zone":
                    this_routing_zone = ws.cell(row=row_num, column=headers[header]).value                
                elif header == "virtual_network":
                    my_id = ws.cell(row=row_num, column=headers[header]).value                
                else:
                    this_virtual_network[header] = ws.cell(row=row_num, column=headers[header]).value
            # skip empty row
            if my_id:
                self.inventory["blueprints"][this_blueprint]["routing_zones"][this_routing_zone]["virtual_networks"][my_id]= this_virtual_network


    def read_systems(self):
        sheet_name = "systems"
        expected = {'blueprint': 1, "system": 2, "role": 3 }
        (ws, headers) = self.read_header(sheet_name, expected)

        for row_num in range(2, ws.max_row+1):
            my_id = ''
            this_blueprint = ''
            this_system = {"interfaces": {}}
            for header in headers:
                if header == "blueprint":
                    this_blueprint = ws.cell(row=row_num, column=headers[header]).value                
                elif header == "system":
                    my_id = ws.cell(row=row_num, column=headers[header]).value                
                else:
                    this_system[header] = ws.cell(row=row_num, column=headers[header]).value
            # skip empty row
            if my_id:
                self.inventory["blueprints"][this_blueprint]["systems"][my_id]= this_system


    def read_interfaces(self):
        sheet_name = "interfaces"
        expected = {'blueprint': 1, "system": 2, "interface": 3, "vns": 4 }
        (ws, headers) = self.read_header(sheet_name, expected)

        for row_num in range(2, ws.max_row+1):
            my_id = ''
            this_blueprint = ''
            this_system = ''
            this_interface = {}
            for header in headers:
                if header == "blueprint":
                    this_blueprint = ws.cell(row=row_num, column=headers[header]).value                
                elif header == "system":
                    this_system = ws.cell(row=row_num, column=headers[header]).value                
                elif header == "interface":
                    my_id = ws.cell(row=row_num, column=headers[header]).value
                else:
                    this_interface[header] = ws.cell(row=row_num, column=headers[header]).value
            # skip empty row
            if my_id:
                self.inventory["blueprints"][this_blueprint]["systems"][this_system]["interfaces"][my_id]= this_interface


    def write_yaml(self):
        print(f"==== final inventory:\n {self.inventory}")

        import yaml
        print( f"==== resulting yaml:\n {yaml.dump(self.inventory)}")

    def run_all(self):
        self.read_aos_server()
        self.read_blueprints()
        self.read_routing_zones()
        self.read_virtual_networks()
        self.read_systems()
        self.read_interfaces()
        # self.write_yaml()


if __name__ == "__main__":
    wb = AosXlsx("fabric-1.xlsx", "temp/temp.yaml")
    wb.run_all()





